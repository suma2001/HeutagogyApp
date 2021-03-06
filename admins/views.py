from django.shortcuts import render, redirect
from django.contrib import messages, auth
import requests

import firebase_admin
import json
from firebase_admin import credentials, firestore 
import pyrebase
import xlrd
import openpyxl

# Create your views here.
firebaseConfig = {
    "apiKey": "AIzaSyAv2Fdg7D52VvYR9TK_H_0z8NiCDj6iQkU",
    "authDomain": "heutagogy-2020.firebaseapp.com",
    "databaseURL": "https://heutagogy-2020.firebaseio.com",
    "projectId": "heutagogy-2020",
    "storageBucket": "heutagogy-2020.appspot.com",
    "messagingSenderId": "772643974559",
    "appId": "1:772643974559:web:d63b880a446d8a70dea1aa",
    "measurementId": "G-X01CP2KNV6"
  }

# cred = credentials.Certificate('heutagogy-2020-6959a4a76c88.json')
firebase = pyrebase.initialize_app(firebaseConfig) 
authe = firebase.auth()
db = firestore.client()

teachers_collection = db.collection('Schools').document('School 1').collection('Teachers')
students_collection = db.collection('Schools').document('School 1').collection('Students') 
admins_collection = db.collection('Schools').document('School 1').collection('Admins')

def admin_dashboard(request):
    if authe.current_user!=None:
        user = authe.current_user
        print(user)
        uid = user['localId']
        results = admins_collection.where('uid', '==', uid).get()[0].to_dict()
        context = {'fullame': results['fullname'], 'email': results['email'], 'photo_url': results['photo_url'], 'dashboard_active': 'active' }
        return render(request, 'admins/admin_dashboard.html', context)
    return redirect('admins:admin_signin')

def admin_signin(request):
    if request.method=="POST":
        ## Get the form data
        data = request.POST.dict()
        email = data.get('emailaddress')
        password = data.get('password')
        print(email, password, "Admins")
        
        try:
            user = authe.sign_in_with_email_and_password(email, password)
            uid = user['localId']
            print(user)
            results = admins_collection.where('uid', '==', uid).get()[0].to_dict()
            print(results)
            messages.success(request, 'Successfully logged in ' + results['fullname'])
            return redirect('admins:admin_dashboard')   
        except requests.exceptions.HTTPError as e:
            error_json = e.args[1]
            error = json.loads(error_json)['error']['message']
            print(error)
            if str(error) == "INVALID_PASSWORD":
                messages.warning(request, 'Invalid login credentials')
            elif str(error) == "EMAIL_NOT_FOUND":
                messages.warning(request, 'Email not found')
            elif str(error) == "TOO_MANY_ATTEMPTS_TRY_LATER : Access to this account has been temporarily disabled due to many failed login attempts. You can immediately restore it by resetting your password or you can try again later.":
                messages.warning(request, 'TOO_MANY_ATTEMPTS_TRY_LATER : Access to this account has been temporarily disabled due to many failed login attempts.')
            return render(request, 'admins/admin_sign_in.html')
    return render(request, 'admins/admin_sign_in.html')

def admin_signup(request):
    if request.method=="POST":
        ## Get the form data
        data = request.POST.dict()
        fullname = data.get('fullname')
        email = data.get('emailaddress')
        password = data.get('password')
        school = data.get('school')
        photo_url = data.get('url')
        
        ## Create a new user
        try:
            user = authe.create_user_with_email_and_password(email, password)
        
            ## Get the UID of the user
            uid = user['localId']
        
            ## Push this user's data to the database
            doc_ref = admins_collection.document(uid)
            doc_ref.set({
                'uid': uid,
                'email': email,
                'fullname': fullname,
                'school':school,
                'photo_url': photo_url,
            })
            user = authe.sign_in_with_email_and_password(email, password)
            return redirect('admins:admin_dashboard')
        except requests.exceptions.HTTPError as e:
            error_json = e.args[1]
            error = json.loads(error_json)['error']['message']
            print(error)
            if str(error) == "EMAIL_EXISTS":
                messages.warning(request, 'Email already exists')
            elif str(error) == "WEAK_PASSWORD : Password should be at least 6 characters":
                messages.warning(request, 'Weak Password : Password should be at least 6 characters')
            return render(request, 'admins/admin_sign_up.html')
        session_id = user['idToken']
        request.session['uid'] = str(session_id)

    return render(request, 'admins/admin_sign_up.html')

def admin_signout(request):
    auth.logout(request)
    authe.current_user=None
    messages.success(request, 'Successfully logged out !')
    return render(request, 'admins/admin_sign_in.html')

def admin_students(request):
    if authe.current_user!=None:
        students = students_collection.stream()
        # print(docs)
        student_list=[]
        i=1
        for doc in students:
            t = (i, doc.to_dict())
            student_list.append(t)
            i+=1
        uid = authe.current_user['localId']
        results = admins_collection.where('uid', '==', uid).get()[0].to_dict()
        context = {'student_list' : student_list, 'fullame': results['fullname'], 'email': results['email'], 'photo_url': results['photo_url'], 'student_active': 'active' }
        return render(request, 'admins/student_list.html', context)
    return redirect('admins:admin_signin')

def admin_teachers(request):
    if authe.current_user!=None:
        teachers = teachers_collection.stream()
        # print(docs)
        teacher_list=[]
        i=1
        for doc in teachers:
            t = (i, doc.to_dict())
            teacher_list.append(t)
            i+=1
        uid = authe.current_user['localId']
        results = admins_collection.where('uid', '==', uid).get()[0].to_dict()
        context = {'teacher_list' : teacher_list, 'fullame': results['fullname'], 'email': results['email'], 'photo_url': results['photo_url'], 'teacher_active': 'active' }
        return render(request, 'admins/teacher_list.html', context)
        
    return redirect('admins:admin_signin')

def find_cols(sheet):
    for col in sheet.iter_rows():
        col_length = len(col)
        break
    print("Col length : ", col_length)
    return col_length

def find_rows(sheet):
    for row in sheet.iter_cols():
        row_length = len(row)
        break
    print("Col length : ", row_length)
    return row_length

def upload_students(request):
    if authe.current_user!=None:
        if request.method=="POST":
            excel_file = request.FILES["excel_file"]
            print(excel_file)
            # loc = ("student_list.xlsx")
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb["Sheet1"]
            print(sheet)

            col_length = find_cols(sheet)
            row_length = find_rows(sheet)
            
            for i in range(2, row_length+1):
                st = "A"+str(i)
                print(st)
                l=[]
                doc_ref = students_collection.document(sheet[st].value)
                doc = doc_ref.get()
                if doc.exists:
                    continue
                else:
                    for j in range(0, col_length):
                        st = chr(ord("A") + j) + str(i)
                        l.append(sheet[st].value)
                    print(l)
                    doc_ref.set({
                        'Roll_No': l[0],
                        'Password': l[1],
                        'Name': l[2],
                        'Class': l[3],
                        'Profile': "",
                        'First_time': True,
                    })
            messages.success(request, 'File uploaded successfully.')
            return redirect('admins:admin_students')
        return render(request, 'admins/add_students.html')
    return redirect('admins:admin_signin')

def upload_teachers(request):
    if authe.current_user!=None:
        if request.method=="POST":
            excel_file = request.FILES["excel_file"]
            print(excel_file)
            # loc = ("student_list.xlsx")
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb["Sheet2"]
            print(sheet)

            col_length = find_cols(sheet)
            row_length = find_rows(sheet)
            
            for i in range(2, row_length+1):
                st = "B"+str(i)
                print(st)
                l=[]
                doc_ref = teachers_collection.document(sheet[st].value)
                doc = doc_ref.get()
                if doc.exists:
                    continue
                else:
                    for j in range(0, col_length):
                        st = chr(ord("A") + j) + str(i)
                        l.append(sheet[st].value)
                    print(l)
                    doc_ref.set({
                        'Name': l[0],
                        'Email': l[1],
                        'courses': [],
                        'Profile': "",
                        'First_time': True,
                    })
            messages.success(request, 'File uploaded successfully.')
            return redirect('admins:admin_teachers')
        return render(request, 'admins/add_teachers.html')
    return redirect('admins:admin_signin')

def add_new_student(request):
    if authe.current_user!=None:
        if request.method=="POST":
            data = request.POST.dict()
            rollno = data.get('rollno')
            password = data.get('password')
            name = data.get('name')
            std = data.get('std')

            doc_ref = students_collection.document(rollno)
            doc_ref.set({
                'Roll_No': rollno,
                'Password': password,
                'Name': name,
                'Class': std,
                'Profile': "",
                'First_time': True,
            })
            messages.success(request, 'Student added successfully.')
            return redirect('admins:admin_students')
        return render(request, 'admins/add_student.html')
    return redirect('admins:admin_signin')


def add_new_teacher(request):
    if authe.current_user!=None:
        if request.method=="POST":
            data = request.POST.dict()
            email = data.get('email')
            password = data.get('password')
            name = data.get('name')

            doc_ref = teachers_collection.document(email)
            doc_ref.set({
                    'Name': name,
                    'Email': email,
                    'courses': [],
                    'Profile': "",
                    'First_time': True,
                })
            messages.success(request, 'Teacher added successfully.')
            return redirect('admins:admin_teachers')
        return render(request, 'admins/add_teacher.html')
    return redirect('admins:admin_signin')
